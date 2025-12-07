# -*- coding: utf-8 -*-
import pandas as pd
import random
import os

# قراءة المنتجات
print("جاري قراءة المنتجات...")
products_df = pd.read_excel('products.xlsx')
products_df.fillna('', inplace=True)
print(f"تم تحميل {len(products_df)} منتج")

# قراءة النوايا الموجودة
existing_intents = []
if os.path.exists('intents.csv'):
    print("جاري قراءة النوايا الموجودة من CSV...")
    existing_df = pd.read_csv('intents.csv', encoding='utf-8-sig')
    existing_intents = existing_df.to_dict('records')
    print(f"تم تحميل {len(existing_intents)} نية موجودة")
else:
    print("لا توجد نوايا موجودة، سيتم إنشاء 1,000,000 نية جديدة")

target_total = 1000000
remaining_needed = target_total - len(existing_intents)
print(f"نحتاج إلى إنشاء {remaining_needed} نية إضافية للوصول إلى 1,000,000")

# قوالب الأسئلة والأجوبة بالعربية المصرية
question_templates = [
    ("كم سعر {product}؟", "سعر {product} هو {price} جنيه. منتج رائع وجدير بالثقة!"),
    ("بكام {product}؟", "متاح بـ {price} جنيه. سعر ممتاز لقيمة عالية!"),
    ("عايز أعرف سعر {product}", "سعر {product} هو {price} جنيه. جودة عالية وسعر مناسب!"),
    ("{product} بكام؟", "{product} بـ {price} جنيه. خيار ممتاز!"),
    ("كم تكلفة {product}؟", "تكلفة {product} هي {price} جنيه. استثمار جيد!"),
    ("{product} متوفر؟", "نعم، {product} متوفر عندنا! جاهز للشحن فوراً."),
    ("عندكم {product}؟", "أكيد! {product} موجود ومتوفر. ممكن نوصلهولك في أقرب وقت."),
    ("{product} موجود؟", "نعم موجود! {product} جاهز للطلب. هنساعدك في أي وقت."),
    ("هل {product} متاح؟", "ممتاز! {product} متاح ومتوفر. جودة عالية وضمان."),
    ("عايز أعرف عن {product}", "{product} منتج ممتاز! {description} جودة عالية وموثوقة."),
    ("إيه مواصفات {product}؟", "{product} يتميز بـ {description} منتج احترافي ومضمون."),
    ("إيه مميزات {product}؟", "{product} فيه مميزات كتير: {description} خيار مثالي!"),
    ("{product} إيه مواصفاته؟", "مواصفات {product}: {description} منتج عالي الجودة!"),
    ("{product} في إيه ألوان؟", "{product} متوفر باللون {color}. لون أنيق وعصري!"),
    ("{product} في لون {color}؟", "نعم! {product} متوفر باللون {color}. لون جميل ومميز!"),
    ("عايز {product} لون {color}", "ممتاز! {product} باللون {color} متوفر. اختيار رائع!"),
    ("عندكم منتجات {brand}؟", "نعم! عندنا منتجات {brand} عالية الجودة. ماركة موثوقة ومشهورة."),
    ("{brand} موجود؟", "أكيد! منتجات {brand} متوفرة. جودة ممتازة وضمان."),
    ("إيه منتجات {brand} المتاحة؟", "عندنا تشكيلة كبيرة من منتجات {brand}. كلها جودة عالية!"),
    ("عندكم {category}؟", "نعم! عندنا تشكيلة ممتازة من {category}. جودة عالية وأسعار مناسبة."),
    ("عايز {category} جيد", "عندنا {category} ممتاز! جودة عالية وموثوقة. هنساعدك تختار الأنسب."),
    ("إيه {category} المتاحة؟", "عندنا تشكيلة واسعة من {category}. كلها منتجات أصلية وموثوقة."),
    ("عايز أشوف {product}", "ممتاز! {product} منتج رائع. {description} جودة عالية وسعر مناسب."),
    ("{product} إيه رأيك فيه؟", "{product} منتج ممتاز! {description} خيار موثوق ومضمون."),
    ("{product} يستاهل؟", "أكيد يستاهل! {product} جودة عالية وسعر مناسب. استثمار جيد."),
    ("{product} كويس؟", "كويس جداً! {product} منتج موثوق. {description} خيار ممتاز."),
    ("عايز أشتري {product}", "ممتاز! {product} خيار رائع. {description} جودة عالية وضمان."),
    ("{product} فيه ضمان؟", "أكيد! {product} فيه ضمان كامل. منتج موثوق ومضمون."),
    ("{product} متوفر للشحن؟", "نعم! {product} متوفر للشحن فوراً. هنسلمك إياه في أقرب وقت."),
    ("إيه أحسن {category} عندكم؟", "عندنا تشكيلة ممتازة من {category}. كلها جودة عالية وأسعار مناسبة."),
    ("عايز {product} بسعر كويس", "{product} سعره {price} جنيه. سعر ممتاز لقيمة عالية!"),
    ("{product} فيه خصم؟", "نعم! {product} متوفر بأسعار ممتازة. {price} جنيه فقط!"),
]

def get_price_range(price):
    if pd.isna(price) or price == '':
        return 'unknown', None, None
    try:
        price_val = float(price)
        if price_val < 500:
            return 'low', 0, 500
        elif price_val < 1500:
            return 'medium', 500, 1500
        elif price_val < 3000:
            return 'high', 1500, 3000
        else:
            return 'premium', 3000, None
    except:
        return 'unknown', None, None

def generate_keywords(title, category, brand, color):
    keywords = []
    if title:
        words = str(title).split()
        keywords.extend([w.strip() for w in words if len(w.strip()) > 2][:5])
    if category:
        keywords.append(str(category).strip())
    if brand:
        keywords.append(str(brand).strip())
    if color:
        keywords.append(str(color).strip())
    return ','.join(keywords[:10])

def generate_patterns(template, title):
    patterns = [template.format(product=title)]
    if len(title.split()) > 1:
        words = title.split()
        short_name = ' '.join(words[:2])
        patterns.append(template.format(product=short_name))
    return ','.join(patterns[:3])

# إنشاء النوايا الإضافية
print(f"جاري إنشاء {remaining_needed} نية إضافية...")
new_intents = []
batch_size = 50000
intents_per_product = max(1, remaining_needed // len(products_df))

for idx, row in products_df.iterrows():
    if len(new_intents) >= remaining_needed:
        break
        
    product_id = int(row['id']) if pd.notna(row['id']) else idx + 1
    title = str(row['title']).strip()
    price = row['price'] if pd.notna(row['price']) else ''
    category = str(row['category']).strip() if pd.notna(row['category']) else ''
    brand = str(row['brand']).strip() if pd.notna(row['brand']) else ''
    color = str(row['color']).strip() if pd.notna(row['color']) else ''
    description = str(row['description']).strip() if pd.notna(row['description']) else 'منتج عالي الجودة'
    
    if not title:
        continue
    
    price_range, min_price, max_price = get_price_range(price)
    price_str = f"{price:.2f}" if pd.notna(price) and price != '' else "غير محدد"
    
    num_intents = random.randint(intents_per_product - 5, intents_per_product + 10)
    
    for _ in range(num_intents):
        if len(new_intents) >= remaining_needed:
            break
            
        q_template, a_template = random.choice(question_templates)
        
        try:
            question = q_template.format(
                product=title,
                color=color if color else "متعدد",
                brand=brand if brand else "متنوع",
                category=category if category else "عام"
            )
            
            response = a_template.format(
                product=title,
                price=price_str,
                description=description if description else "منتج عالي الجودة وموثوق",
                color=color if color else "متعدد الألوان",
                brand=brand if brand else "ماركات مختلفة",
                category=category if category else "منتجات متنوعة"
            )
            
            intent_name = f"intent_{product_id}_{random.randint(10000, 99999)}"
            keywords = generate_keywords(title, category, brand, color)
            patterns = generate_patterns(q_template, title)
            
            new_intents.append({
                'intent': intent_name,
                'patterns': patterns,
                'keywords': keywords,
                'response': response,
                'color': color.lower() if color else '',
                'price_range': price_range,
                'min_price': min_price if min_price is not None else '',
                'max_price': max_price if max_price is not None else '',
                'category': category.lower() if category else '',
                'product_ids': str(product_id)
            })
        except:
            continue
    
    if (idx + 1) % 1000 == 0:
        print(f"تم إنشاء {len(new_intents)} نية جديدة... ({len(new_intents) + len(existing_intents)} إجمالي)")

# إضافة المزيد إذا لزم
print("جاري إضافة المزيد للوصول إلى 1,000,000...")
while len(new_intents) < remaining_needed:
    row = products_df.iloc[random.randint(0, len(products_df) - 1)]
    product_id = int(row['id']) if pd.notna(row['id']) else random.randint(1, 10000)
    title = str(row['title']).strip()
    price = row['price'] if pd.notna(row['price']) else ''
    category = str(row['category']).strip() if pd.notna(row['category']) else ''
    brand = str(row['brand']).strip() if pd.notna(row['brand']) else ''
    color = str(row['color']).strip() if pd.notna(row['color']) else ''
    description = str(row['description']).strip() if pd.notna(row['description']) else 'منتج عالي الجودة'
    
    if not title:
        continue
    
    q_template, a_template = random.choice(question_templates)
    price_range, min_price, max_price = get_price_range(price)
    price_str = f"{price:.2f}" if pd.notna(price) and price != '' else "غير محدد"
    
    try:
        question = q_template.format(
            product=title,
            color=color if color else "متعدد",
            brand=brand if brand else "متنوع",
            category=category if category else "عام"
        )
        
        response = a_template.format(
            product=title,
            price=price_str,
            description=description if description else "منتج عالي الجودة وموثوق",
            color=color if color else "متعدد الألوان",
            brand=brand if brand else "ماركات مختلفة",
            category=category if category else "منتجات متنوعة"
        )
        
        intent_name = f"intent_{product_id}_{random.randint(10000, 99999)}"
        keywords = generate_keywords(title, category, brand, color)
        patterns = generate_patterns(q_template, title)
        
        new_intents.append({
            'intent': intent_name,
            'patterns': patterns,
            'keywords': keywords,
            'response': response,
            'color': color.lower() if color else '',
            'price_range': price_range,
            'min_price': min_price if min_price is not None else '',
            'max_price': max_price if max_price is not None else '',
            'category': category.lower() if category else '',
            'product_ids': str(product_id)
        })
    except:
        continue

new_intents = new_intents[:remaining_needed]
print(f"تم إنشاء {len(new_intents)} نية جديدة بنجاح!")

# دمج النوايا
all_intents = existing_intents + new_intents
print(f"إجمالي النوايا: {len(all_intents)}")

# حفظ في Excel على دفعات
print("جاري حفظ الملف في Excel...")
try:
    # استخدام xlsxwriter للكفاءة
    writer = pd.ExcelWriter('intents.xlsx', engine='xlsxwriter', options={'strings_to_urls': False})
    
    # حفظ على دفعات لتوفير الذاكرة
    chunk_size = 100000
    for i in range(0, len(all_intents), chunk_size):
        chunk = all_intents[i:i+chunk_size]
        df_chunk = pd.DataFrame(chunk)
        
        if i == 0:
            df_chunk.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0)
        else:
            df_chunk.to_excel(writer, index=False, sheet_name='Sheet1', startrow=i, header=False)
        
        print(f"تم حفظ {min(i+chunk_size, len(all_intents))} من {len(all_intents)} صف...")
    
    writer.close()
    print(f"✓✓✓ تم حفظ {len(all_intents)} نية في ملف intents.xlsx بنجاح!")
    
except Exception as e:
    print(f"خطأ مع xlsxwriter: {e}")
    print("جاري المحاولة مع openpyxl...")
    try:
        # حفظ على دفعات أصغر
        chunk_size = 50000
        first_chunk = True
        
        for i in range(0, len(all_intents), chunk_size):
            chunk = all_intents[i:i+chunk_size]
            df_chunk = pd.DataFrame(chunk)
            
            if first_chunk:
                df_chunk.to_excel('intents.xlsx', index=False, engine='openpyxl')
                first_chunk = False
            else:
                from openpyxl import load_workbook
                book = load_workbook('intents.xlsx')
                writer = pd.ExcelWriter('intents.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
                writer.book = book
                df_chunk.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
                writer.close()
            
            print(f"تم حفظ {min(i+chunk_size, len(all_intents))} من {len(all_intents)} صف...")
        
        print(f"✓✓✓ تم حفظ {len(all_intents)} نية في ملف intents.xlsx بنجاح!")
    except Exception as e2:
        print(f"خطأ في الحفظ: {e2}")
        print("جاري الحفظ في CSV كبديل...")
        df_all = pd.DataFrame(all_intents)
        df_all.to_csv('intents.csv', index=False, encoding='utf-8-sig')
        print(f"تم حفظ {len(all_intents)} نية في ملف intents.csv")
        print("يمكنك فتحه في Excel وتحويله يدوياً")

print("\nالملف جاهز للاستخدام!")

